#!/usr/bin/env python3
"""
clean_to_schema.py

Purpose:
    Clean exclusion-list CSV/XLSX files and rewrite them to match the updated
    staging-table schema:
      - id
      - all schema columns as VARCHAR-compatible text
      - required workflow/source columns filled; sparse source fields can stay blank
      - business_name included instead of relying only on entity_name
      - dates normalized to YYYY-MM-DD when possible, but still written as text

How to run from PowerShell:
    python clean_to_schema.py --input-dir . --output-dir new_schema_ready_csvs

Typical use:
    1. Put your processed files in one folder, for example:
       stg_arizona_exclusions.xlsx
       stg_arkansas.exclusions.csv
       stg_california_exclusions_processed.csv
       etc.
    2. Run this script from that folder.
    3. Import the output CSVs into the matching PostgreSQL staging tables.
       Output files are named stg_<state>_exclusions.csv.

No database connection is used. This only cleans source files.
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

NEW_COLUMNS: List[str] = [
    "id", "record_type", "source_state", "source_state_abbr", "source_name",
    "provider_name", "first_name", "middle_name", "last_name", "business_name",
    "aka", "dba", "npi", "provider_type", "license_number", "provider_number",
    "action_type", "action_effective_date", "active_period", "exclusion_authority",
    "exclusion_reason", "reinstatement_date", "source_url", "source_file_url",
    "source_file_date", "date_accessed", "data_quality_status", "notes",
]

NULL_LIKE = {"", "null", "none", "nan", "na", "n/a", "[null]", "nat"}
ORG_MARKERS = [" llc", " inc", " corp", " company", " clinic", " center", " services", " group", " dba ", " pc", " ltd"]
CREDENTIAL_SUFFIXES = {
    "MD", "DO", "DPM", "DMD", "DDS", "RN", "LPN", "NP", "PA", "PHD",
    "PHARMD", "DC", "OD", "MSW", "LCSW", "CNA",
}

DEFAULT_BLANK_COLUMNS: Set[str] = {
    "first_name",
    "middle_name",
    "last_name",
    "business_name",
    "aka",
    "dba",
    "npi",
    "provider_type",
    "license_number",
    "provider_number",
    "action_type",
    "action_effective_date",
    "active_period",
    "exclusion_authority",
    "exclusion_reason",
    "reinstatement_date",
    "source_url",
    "source_file_url",
    "source_file_date",
    "date_accessed",
    "notes",
}

STATE_FROM_FILE = {
    "alabama": ("Alabama", "AL"),
    "alaska": ("Alaska", "AK"),
    "arizona": ("Arizona", "AZ"),
    "arkansas": ("Arkansas", "AR"),
    "california": ("California", "CA"),
    "colorado": ("Colorado", "CO"),
    "connecticut": ("Connecticut", "CT"),
    "delaware": ("Delaware", "DE"),
    "district_of_columbia": ("District of Columbia", "DC"),
    "dc": ("District of Columbia", "DC"),
    "florida": ("Florida", "FL"),
}

STATE_BY_ABBR = {
    state_abbr.lower(): (state_name, state_abbr)
    for state_name, state_abbr in set(STATE_FROM_FILE.values())
}
STATE_BY_NORMALIZED_NAME = {
    re.sub(r"[^a-z]+", " ", state_name.lower()).strip(): (state_name, state_abbr)
    for state_name, state_abbr in set(STATE_FROM_FILE.values())
}
EXPLICIT_STATE_COLUMNS = {
    "source state",
    "source state abbr",
    "source state abbrev",
    "state abbr",
    "state abbrev",
    "state abbreviation",
}
GENERIC_STATE_COLUMNS = {"state"}

DATE_FORMATS = [
    "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
    "%B %d, %Y", "%b %d, %Y", "%Y%m%d",
]

SYNONYMS = {
    "source_state_abbr": ["source_state_abbr", "source_state_abbrev", "state_abbr"],
    "provider_name": ["provider_name", "name", "provider", "provider name", "name of provider", "sanctioned provider name", "name of individual", "name of company"],
    "first_name": ["first_name", "firstname", "first name"],
    "middle_name": ["middle_name", "midname", "middle name", "mi"],
    "last_name": ["last_name", "lastname", "last name"],
    "business_name": ["business_name", "entity_name", "busname", "business", "facility name", "name of company"],
    "aka": ["aka", "a/k/a", "alias"],
    "dba": ["dba", "doing business as", "business_name"],
    "npi": ["npi", "npi #", "npi number"],
    "provider_type": ["provider_type", "provider type", "specialty", "general"],
    "license_number": ["license_number", "license number", "license #", "licnum"],
    "provider_number": ["provider_number", "provider number", "upin", "dea #", "dea number"],
    "action_type": ["action_type", "administrative action", "oig / medicaid sanction", "excltype"],
    "action_effective_date": ["action_effective_date", "effective date", "exclusion effective date", "date of suspension", "suspension effective date", "termination effective date", "action date", "excldate"],
    "active_period": ["active_period", "period", "active period"],
    "exclusion_authority": ["exclusion_authority", "termination authority", "agency instituting the action", "suspension initiated by", "initiated by", "oig / medicaid sanction", "excltype"],
    "exclusion_reason": ["exclusion_reason", "reason for the action", "administrative action", "oig / medicaid sanction", "excltype"],
    "reinstatement_date": ["reinstatement_date", "reinstated date", "exclusion end date", "expiration date", "termination date", "termination date﻿", "reindate"],
    "source_url": ["source_url"],
    "source_file_url": ["source_file_url"],
    "source_file_date": ["source_file_date"],
    "date_accessed": ["date_accessed"],
    "data_quality_status": ["data_quality_status"],
    "notes": ["notes", "comments", "address", "principal address", "address(es)"],
}


def normalize_header(value: str) -> str:
    value = value.replace("\ufeff", "").strip().lower()
    value = re.sub(r"\s+", " ", value)
    return value


HEADER_ALIASES = {
    normalize_header(alias)
    for canonical, aliases in SYNONYMS.items()
    for alias in [canonical, *aliases]
}


def clean_text(value: object, default: str = "N/A") -> str:
    if value is None:
        return default
    text = str(value).replace("\ufeff", "").replace("\r", " ").replace("\n", " ").replace("\t", " ").strip()
    text = re.sub(r"\s+", " ", text)
    if text.lower() in NULL_LIKE:
        return default
    return text


def has_value(value: object) -> bool:
    return clean_text(value, default="") != ""


def get_value(row: Dict[str, str], canonical: str, default: str = "N/A") -> str:
    candidates = SYNONYMS.get(canonical, [canonical])
    for candidate in candidates:
        key = normalize_header(candidate)
        if key in row:
            value = clean_text(row.get(key), default="")
            if value and value.lower() not in NULL_LIKE:
                return value
    return default


def infer_state_from_filename(path: Path) -> Tuple[str, str]:
    stem = path.stem.lower()
    for token, value in STATE_FROM_FILE.items():
        if token in stem:
            return value
    return "N/A", "N/A"


def state_slug(state_name: str) -> str:
    return re.sub(r"_+", "_", normalize_header(state_name).replace(" ", "_"))


def normalize_state_value(value: object) -> str:
    text = clean_text(value, default="")
    text = re.sub(r"[^A-Za-z]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def normalize_state_column(value: object) -> str:
    return normalize_state_value(str(value).replace("_", " "))


def state_from_exact_value(value: object, allow_abbr: bool = True) -> Optional[Tuple[str, str]]:
    token = normalize_state_value(value)
    if not token:
        return None
    if token in STATE_BY_NORMALIZED_NAME:
        return STATE_BY_NORMALIZED_NAME[token]
    compact = token.replace(" ", "")
    if allow_abbr and compact in STATE_BY_ABBR:
        return STATE_BY_ABBR[compact]
    return None


def state_from_text(value: object) -> Optional[Tuple[str, str]]:
    text = normalize_state_value(value)
    if not text:
        return None

    matches = []
    for normalized_name, state_info in STATE_BY_NORMALIZED_NAME.items():
        if re.search(rf"(^| ){re.escape(normalized_name)}( |$)", text):
            matches.append(state_info)

    unique_matches = sorted(set(matches))
    if len(unique_matches) == 1:
        return unique_matches[0]
    return None


def top_state_vote(votes: Dict[Tuple[str, str], int]) -> Optional[Tuple[Tuple[str, str], int]]:
    if not votes:
        return None
    return max(votes.items(), key=lambda item: item[1])


def infer_state_from_content(headers: List[str], rows: List[Dict[str, str]]) -> Optional[Tuple[str, str]]:
    explicit_votes: Dict[Tuple[str, str], int] = {}
    generic_votes: Dict[Tuple[str, str], int] = {}
    generic_values = 0

    for row in rows:
        for key, value in row.items():
            column = normalize_state_column(key)
            if column in EXPLICIT_STATE_COLUMNS:
                state_info = state_from_exact_value(value, allow_abbr=True)
                if state_info:
                    explicit_votes[state_info] = explicit_votes.get(state_info, 0) + 1
            elif column in GENERIC_STATE_COLUMNS:
                if has_value(value):
                    generic_values += 1
                state_info = state_from_exact_value(value, allow_abbr=True)
                if state_info:
                    generic_votes[state_info] = generic_votes.get(state_info, 0) + 1

    explicit_top = top_state_vote(explicit_votes)
    if explicit_top:
        return explicit_top[0]

    generic_top = top_state_vote(generic_votes)
    if generic_top:
        _, top_count = generic_top
        if generic_values <= 1 or top_count / max(generic_values, 1) >= 0.8:
            return generic_top[0]

    text_votes: Dict[Tuple[str, str], int] = {}
    for header in headers:
        state_info = state_from_text(header)
        if state_info:
            text_votes[state_info] = text_votes.get(state_info, 0) + 1
    for row in rows:
        for value in row.values():
            state_info = state_from_text(value)
            if state_info:
                text_votes[state_info] = text_votes.get(state_info, 0) + 1

    text_top = top_state_vote(text_votes)
    if text_top:
        top_state, top_count = text_top
        tied = [state_info for state_info, count in text_votes.items() if count == top_count]
        if len(tied) == 1:
            return top_state

    return None


def infer_state_from_file_strings(path: Path) -> Optional[Tuple[str, str]]:
    text_votes: Dict[Tuple[str, str], int] = {}

    def add_vote(value: object) -> None:
        state_info = state_from_text(value)
        if state_info:
            text_votes[state_info] = text_votes.get(state_info, 0) + 1

    suffix = path.suffix.lower()
    if suffix == ".csv":
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                with path.open("r", encoding=encoding, errors="strict") as f:
                    for _ in range(200):
                        line = f.readline()
                        if not line:
                            break
                        add_vote(line)
                break
            except UnicodeDecodeError:
                continue
    elif suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row_index, raw_row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_index >= 200:
                        break
                    for value in raw_row:
                        if value is not None:
                            add_vote(excel_value_to_text(value))
        finally:
            workbook.close()

    text_top = top_state_vote(text_votes)
    if not text_top:
        return None

    top_state, top_count = text_top
    tied = [state_info for state_info, count in text_votes.items() if count == top_count]
    if len(tied) == 1:
        return top_state
    return None


def infer_state(path: Path, headers: List[str], rows: List[Dict[str, str]]) -> Tuple[str, str]:
    return (
        infer_state_from_content(headers, rows)
        or infer_state_from_file_strings(path)
        or infer_state_from_filename(path)
    )


def normalize_date(value: object, default: str = "N/A") -> str:
    text = clean_text(value, default=default)
    if text in {"", "N/A"}:
        return text
    if text.lower() in {"indefinite", "ongoing", "current", "active"}:
        return text
    if re.fullmatch(r"\d+(\.0+)?", text):
        serial = float(text)
        if 20000 <= serial <= 60000:
            parsed = datetime(1899, 12, 30) + timedelta(days=serial)
            return parsed.strftime("%Y-%m-%d")
    candidate = re.sub(r"\b(\d{1,2})(st|nd|rd|th)\b", r"\1", text, flags=re.IGNORECASE)
    candidate = re.sub(r"\s+", " ", candidate.strip())
    for fmt in DATE_FORMATS:
        try:
            parsed = datetime.strptime(candidate, fmt)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def is_date_like(value: object) -> bool:
    text = clean_text(value, default="")
    if not text:
        return True
    if text.lower() in {"indefinite", "ongoing", "current", "active"}:
        return True
    if re.fullmatch(r"\d+(\.0+)?", text):
        serial = float(text)
        return 20000 <= serial <= 60000
    candidate = re.sub(r"\b(\d{1,2})(st|nd|rd|th)\b", r"\1", text, flags=re.IGNORECASE)
    candidate = re.sub(r"\s+", " ", candidate.strip())
    for fmt in DATE_FORMATS:
        try:
            datetime.strptime(candidate, fmt)
            return True
        except ValueError:
            continue
    return False


def parse_person_name(provider_name: str) -> Tuple[str, str, str]:
    name = clean_text(provider_name)
    if name == "N/A":
        return "N/A", "N/A", "N/A"
    if any(marker in f" {name.lower()}" for marker in ORG_MARKERS):
        return "N/A", "N/A", "N/A"
    if "," in name:
        last, rest = [part.strip() for part in name.split(",", 1)]
        normalized_rest = re.sub(r"[^A-Za-z]", "", rest).upper()
        if normalized_rest in CREDENTIAL_SUFFIXES:
            return parse_person_name(last)
        pieces = rest.split()
        first = pieces[0] if pieces else "N/A"
        middle = " ".join(pieces[1:]) if len(pieces) > 1 else "N/A"
        return clean_text(first), clean_text(middle), clean_text(last)
    pieces = name.split()
    if len(pieces) >= 2:
        first = pieces[0]
        last = pieces[-1]
        middle = " ".join(pieces[1:-1]) if len(pieces) > 2 else "N/A"
        return clean_text(first), clean_text(middle), clean_text(last)
    return "N/A", "N/A", name


def looks_like_organization_name(provider_name: str) -> bool:
    name = clean_text(provider_name, default="")
    return any(marker in f" {name.lower()}" for marker in ORG_MARKERS)


def normalize_npi(value: object, default: str = "N/A") -> str:
    text = clean_text(value, default=default)
    if text in {"", "N/A"}:
        return text
    if re.fullmatch(r"0+", text):
        return default
    return text


def column_default(column: str, blank_columns: Set[str]) -> str:
    return "" if column in blank_columns else "N/A"


def build_new_row(
    old_row: Dict[str, str],
    row_id: int,
    state_info: Tuple[str, str],
    blank_columns: Set[str],
) -> Dict[str, str]:
    state_name, state_abbr = state_info
    new = {column: column_default(column, blank_columns) for column in NEW_COLUMNS}
    new["id"] = str(row_id)
    new["record_type"] = get_value(old_row, "record_type", "provider_record")
    new["source_state"] = get_value(old_row, "source_state", state_name)
    new["source_state_abbr"] = get_value(old_row, "source_state_abbr", state_abbr)
    new["source_name"] = get_value(old_row, "source_name", f"{new['source_state']} exclusion source")

    for col in NEW_COLUMNS:
        if col in {"id", "record_type", "source_state", "source_state_abbr", "source_name", "business_name"}:
            continue
        default = column_default(col, blank_columns)
        if col in {"action_effective_date", "reinstatement_date", "source_file_date", "date_accessed"}:
            new[col] = normalize_date(get_value(old_row, col, default), default=default)
        elif col == "npi":
            new[col] = normalize_npi(get_value(old_row, col, default), default=default)
        else:
            new[col] = get_value(old_row, col, default)

    new["business_name"] = get_value(old_row, "business_name", column_default("business_name", blank_columns))
    if not has_value(new["provider_name"]):
        if has_value(new["business_name"]):
            new["provider_name"] = new["business_name"]
        else:
            parts = [new["first_name"], new["middle_name"], new["last_name"]]
            parts = [p for p in parts if has_value(p)]
            if parts:
                new["provider_name"] = " ".join(parts)

    parsed_first, parsed_middle, parsed_last = parse_person_name(new["provider_name"])
    if not has_value(new["first_name"]) and has_value(parsed_first):
        new["first_name"] = parsed_first
    if not has_value(new["middle_name"]) and has_value(parsed_middle):
        new["middle_name"] = parsed_middle
    if not has_value(new["last_name"]) and has_value(parsed_last):
        new["last_name"] = parsed_last

    if not has_value(new["business_name"]) and has_value(new["dba"]):
        new["business_name"] = new["dba"]
    if not has_value(new["business_name"]) and looks_like_organization_name(new["provider_name"]):
        new["business_name"] = new["provider_name"]

    for col in NEW_COLUMNS:
        new[col] = clean_text(new[col], default=column_default(col, blank_columns))
    if not has_value(new["data_quality_status"]):
        new["data_quality_status"] = "clean_ready_for_import"
    return new


def read_csv_with_encoding(path: Path, encoding: str):
    with path.open("r", encoding=encoding, newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError(f"No header row found in {path}")
        normalized_headers = [normalize_header(h) for h in reader.fieldnames]
        rows = []
        for raw in reader:
            normalized = {}
            for original, normalized_key in zip(reader.fieldnames, normalized_headers):
                normalized[normalized_key] = raw.get(original, "")
            rows.append(normalized)
        return normalized_headers, rows


def read_csv(path: Path):
    last_error = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return read_csv_with_encoding(path, encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error or ValueError(f"Could not read {path}")


def excel_value_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def combine_header_rows(row: List[str], next_row: List[str] | None = None) -> List[str]:
    width = max(len(row), len(next_row or []))
    headers = []
    for index in range(width):
        value = clean_text(row[index] if index < len(row) else "", default="")
        next_value = clean_text(next_row[index] if next_row and index < len(next_row) else "", default="")
        if value and next_value and normalize_header(value) != normalize_header(next_value):
            headers.append(f"{value} {next_value}")
        else:
            headers.append(value or next_value)
    return headers


def header_score(headers: List[str]) -> int:
    score = 0
    for header in headers:
        normalized = normalize_header(header)
        if not normalized:
            continue
        if normalized in HEADER_ALIASES:
            score += 4
            continue
        if "name" in normalized and any(token in normalized for token in ("provider", "individual", "company")):
            score += 3
        if "effective date" in normalized or "action date" in normalized:
            score += 3
        if "initiated by" in normalized or "authority" in normalized:
            score += 3
        if normalized in {"npi", "license number", "provider number"}:
            score += 3
    return score


def detect_header_row(rows: List[List[str]]) -> Tuple[int, int, List[str]]:
    best_index = -1
    best_span = 1
    best_headers: List[str] = []
    best_score = 0

    for index, row in enumerate(rows):
        one_row_headers = combine_header_rows(row)
        one_row_score = header_score(one_row_headers)
        if one_row_score > best_score:
            best_index = index
            best_span = 1
            best_headers = one_row_headers
            best_score = one_row_score

        if index + 1 < len(rows):
            two_row_headers = combine_header_rows(row, rows[index + 1])
            two_row_score = header_score(two_row_headers)
            if two_row_score > best_score:
                best_index = index
                best_span = 2
                best_headers = two_row_headers
                best_score = two_row_score

    if best_index < 0 or best_score < 6:
        raise ValueError("No recognizable header row found")
    return best_index, best_span, best_headers


def looks_like_section_label(value: object) -> bool:
    text = clean_text(value, default="")
    if not text or len(text) > 80:
        return False
    if "," in text or re.search(r"\d", text):
        return False
    return text.upper() == text


def looks_like_prose_note(value: object) -> bool:
    text = clean_text(value, default="")
    if len(text) > 100 or text.endswith("."):
        return True
    words = text.split()
    if len(words) >= 10 and "," not in text:
        return True
    prose_markers = (" medicaid ", " database ", " website ", " provider ", " providers ")
    return any(marker in f" {text.lower()} " for marker in prose_markers)


def read_xlsx(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read XLSX files. Install requirements.txt first.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            nonempty_rows = []
            row_iter = sheet.iter_rows(values_only=True)
            for raw_row in row_iter:
                values = [excel_value_to_text(value) for value in raw_row]
                if any(clean_text(value, default="") for value in values):
                    nonempty_rows.append(values)
            if not nonempty_rows:
                continue

            header_index, header_span, headers = detect_header_row(nonempty_rows)
            normalized_headers = [normalize_header(header) for header in headers]
            rows = []
            current_provider_type = ""
            for values in nonempty_rows[header_index + header_span:]:
                populated_indexes = [
                    index
                    for index, value in enumerate(values)
                    if clean_text(value, default="")
                ]
                if not populated_indexes:
                    continue
                if (
                    len(populated_indexes) == 1
                    and populated_indexes[0] == 0
                    and looks_like_section_label(values[0])
                ):
                    current_provider_type = clean_text(values[0], default="")
                    continue
                if header_score(combine_header_rows(values)) >= 6:
                    continue
                normalized = {}
                for index, normalized_key in enumerate(normalized_headers):
                    if not normalized_key:
                        continue
                    normalized[normalized_key] = values[index] if index < len(values) else ""
                date_value = get_value(normalized, "action_effective_date", default="")
                if date_value and not is_date_like(date_value):
                    continue
                provider_value = get_value(normalized, "provider_name", default="")
                if not date_value and looks_like_prose_note(provider_value):
                    continue
                if current_provider_type and not get_value(normalized, "provider_type", default=""):
                    normalized["provider_type"] = current_provider_type
                rows.append(normalized)
            return normalized_headers, rows
    finally:
        workbook.close()

    raise ValueError(f"No header row found in {path}")


def read_source_file(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise ValueError(f"Unsupported file type: {path.suffix}")


def should_process(path: Path) -> bool:
    name = path.name.lower()
    if name.startswith("~$"):
        return False
    if path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
        return False
    if "summary" in name or "load_summary" in name or "source_match" in name:
        return False
    return name.startswith("stg_") or name.startswith("all_states")


def output_filename(path: Path, state_info: Tuple[str, str]) -> str:
    state_name, _ = state_info
    if state_name != "N/A":
        return f"stg_{state_slug(state_name)}_exclusions.csv"
    stem = path.stem
    return f"{stem}_schema.csv"


def process_file(path: Path, output_dir: Path, blank_columns: Set[str]) -> Dict[str, str]:
    headers, rows = read_source_file(path)
    state_info = infer_state(path, headers, rows)
    output_path = output_dir / output_filename(path, state_info)
    cleaned_rows = []

    for row in rows:
        cleaned = build_new_row(row, len(cleaned_rows) + 1, state_info, blank_columns)
        cleaned["id"] = str(len(cleaned_rows) + 1)
        cleaned_rows.append(cleaned)

    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=NEW_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(cleaned_rows)

    return {
        "source_file": path.name,
        "output_file": output_path.name,
        "detected_state": state_info[0],
        "raw_rows": str(len(rows)),
        "output_rows": str(len(cleaned_rows)),
        "duplicates_removed": "0",
        "output_columns": str(len(NEW_COLUMNS)),
        "status": "processed",
    }


def parse_blank_columns(value: str) -> Set[str]:
    columns = {normalize_header(item).replace(" ", "_") for item in value.split(",") if item.strip()}
    unknown = sorted(columns - set(NEW_COLUMNS))
    if unknown:
        raise argparse.ArgumentTypeError(f"Unknown column(s): {', '.join(unknown)}")
    return columns


def main() -> None:
    parser = argparse.ArgumentParser(description="Clean exclusion CSV/XLSX files into staging schema.")
    parser.add_argument("--input-dir", default=".", help="Folder containing source CSV/XLSX files.")
    parser.add_argument("--output-dir", default="new_schema_ready_csvs", help="Folder for cleaned output CSVs.")
    parser.add_argument(
        "--blank-columns",
        type=parse_blank_columns,
        default=DEFAULT_BLANK_COLUMNS,
        help=(
            "Comma-separated output columns that may remain blank instead of N/A. "
            f"Default: {','.join(sorted(DEFAULT_BLANK_COLUMNS))}"
        ),
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(path for path in input_dir.iterdir() if path.is_file() and should_process(path))
    if not files:
        raise SystemExit(f"No staging CSV/XLSX files found in {input_dir}")

    summary_rows = []
    for path in files:
        try:
            result = process_file(path, output_dir, args.blank_columns)
        except Exception as exc:
            result = {
                "source_file": path.name,
                "output_file": "N/A",
                "detected_state": "N/A",
                "raw_rows": "N/A",
                "output_rows": "N/A",
                "duplicates_removed": "N/A",
                "output_columns": str(len(NEW_COLUMNS)),
                "status": f"error: {exc}",
            }
        summary_rows.append(result)

    summary_path = output_dir / "schema_cleaning_summary.csv"
    with summary_path.open("w", encoding="utf-8", newline="") as f:
        fieldnames = [
            "source_file",
            "output_file",
            "detected_state",
            "raw_rows",
            "output_rows",
            "duplicates_removed",
            "output_columns",
            "status",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(summary_rows)

    print(f"Processed {len(summary_rows)} file(s).")
    print(f"Output folder: {output_dir}")
    print(f"Summary file: {summary_path}")


if __name__ == "__main__":
    main()
